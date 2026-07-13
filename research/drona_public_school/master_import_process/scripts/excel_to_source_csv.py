#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

from common import (
    SOURCE_FILES,
    applies_token_matches,
    next_academic_year,
    normalise_cell,
    resolve_tokens,
    stream_applies_to_grade,
    subject_applies_to_grade_stream,
)


def row_values(row, width: int) -> list[str]:
    values = [normalise_cell(cell.value) for cell in row[:width]]
    if len(values) < width:
        values.extend([""] * (width - len(values)))
    return values


def sheet_metadata(ws) -> dict[str, str]:
    cells = [normalise_cell(cell.value) for cell in ws[1]]
    metadata: dict[str, str] = {}
    for index in range(0, len(cells) - 1, 2):
        key = cells[index]
        value = cells[index + 1]
        if key:
            metadata[key] = value
    return metadata


def metadata_int(metadata: dict[str, str], key: str, default: int) -> int:
    value = metadata.get(key)
    if not value:
        return default
    try:
        return int(value)
    except ValueError:
        return default


def sheet_to_rows(ws, skip_example_row: bool = False) -> tuple[list[str], list[list[str]]]:
    metadata = sheet_metadata(ws)
    header_row = metadata_int(metadata, "header_row", 2)
    headers = [normalise_cell(cell.value) for cell in ws[header_row] if normalise_cell(cell.value)]
    width = len(headers)
    rows: list[list[str]] = []
    example_row = metadata_int(metadata, "example_row", 0) if skip_example_row else 0
    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        if row_number == example_row:
            continue
        values = row_values(row, width)
        if any(values):
            rows.append(values)
    return headers, rows


def dict_rows(headers: list[str], rows: list[list[str]]) -> list[dict[str, str]]:
    return [dict(zip(headers, row)) for row in rows]


def rows_for(sheets: dict[str, tuple[list[str], list[list[str]]]], sheet: str) -> list[dict[str, str]]:
    headers, rows = sheets.get(sheet, ([], []))
    return dict_rows(headers, rows)


def auto_generate_subject_matrix(sheets: dict[str, tuple[list[str], list[list[str]]]]) -> None:
    headers, rows = sheets.get("09_subject_matrix", ([], []))
    if not headers:
        return
    boards = rows_for(sheets, "03_boards")
    mediums = rows_for(sheets, "04_mediums")
    grades = rows_for(sheets, "05_grades")
    streams = rows_for(sheets, "06_streams")
    subjects = rows_for(sheets, "08_subjects")
    if not all([boards, mediums, grades, streams, subjects]):
        return

    generated: list[list[str]] = []
    for board in boards:
        for medium in mediums:
            for grade in grades:
                grade_code = grade.get("grade_code", "")
                for stream in streams:
                    stream_code = stream.get("stream_code", "")
                    applies_to = stream.get("applies_to", "").strip()
                    if not stream_applies_to_grade(grade_code, stream_code, applies_to):
                        continue
                    for subject in subjects:
                        subject_applies_to = subject.get("applies_to", "") or subject.get("notes", "")
                        if not subject_applies_to_grade_stream(grade_code, stream_code, subject_applies_to):
                            continue
                        row = {
                            "board_code": board.get("board_code", ""),
                            "medium_code": medium.get("medium_code", ""),
                            "grade_code": grade_code,
                            "stream_code": stream_code,
                            "subject_code": subject.get("subject_code", ""),
                            "subject_name": subject.get("subject_name", ""),
                            "subject_category": subject.get("subject_category", ""),
                            "is_compulsory": subject.get("matrix_is_compulsory", ""),
                            "is_elective": subject.get("matrix_is_elective", ""),
                            "display_order": subject.get("matrix_display_order", ""),
                            "source_note": subject.get("matrix_source_note", ""),
                        }
                        generated.append([row.get(header, "") for header in headers])
    sheets["09_subject_matrix"] = (headers, generated)


def workbook_next_year(sheets: dict[str, tuple[list[str], list[list[str]]]], year: str) -> str:
    years = [row.get("academic_year", "") for row in rows_for(sheets, "02_academic_years")]
    if year in years:
        index = years.index(year)
        if index + 1 < len(years):
            return years[index + 1]
    return next_academic_year(year)


def is_alumni_source_cohort(cohort: dict[str, str], rules: list[dict[str, str]]) -> bool:
    grade_code = cohort.get("grade_code", "")
    stream_code = cohort.get("stream_code", "")
    for rule in rules:
        to_grade = rule.get("to_grade_code", "").upper()
        decision = rule.get("promotion_decision", "").upper()
        from_grade = rule.get("from_grade_code", "")
        from_stream = rule.get("from_stream_code", "").upper()
        if to_grade != "ALUMNI" and decision != "ALUMNI":
            continue
        if not applies_token_matches(from_grade, grade_code, stream_code):
            continue
        if from_stream and from_stream not in {"ALL", stream_code.upper()}:
            continue
        return True
    return False


def auto_generate_alumni_cohorts(sheets: dict[str, tuple[list[str], list[list[str]]]], year: str) -> None:
    headers, rows = sheets.get("58_alumni_cohorts", ([], []))
    if not headers:
        return
    cohorts = rows_for(sheets, "14_cohorts")
    promotion_rules = rows_for(sheets, "45_promotion_rules")
    if not cohorts or not promotion_rules:
        return

    alumni_year = workbook_next_year(sheets, year)
    start_year = alumni_year.split("-", 1)[0]
    generated: list[list[str]] = []
    for cohort in cohorts:
        if not is_alumni_source_cohort(cohort, promotion_rules):
            continue
        cohort_code = (
            f"{cohort.get('school_code', '')}-ALUMNI-{start_year}-"
            f"{cohort.get('board_code', '')}-{cohort.get('medium_code', '')}-"
            f"{cohort.get('grade_code', '')}-{cohort.get('stream_code', '')}-"
            f"{cohort.get('division_code', '')}"
        )
        row = {
            **cohort,
            "cohort_code": cohort_code,
            "name": (
                f"{cohort.get('school_code', '')} Alumni {alumni_year} "
                f"{cohort.get('medium_code', '')} {cohort.get('grade_code', '')} "
                f"{cohort.get('stream_code', '')} Division {cohort.get('division_code', '')}"
            ),
            "idnumber": cohort_code,
            "academic_year": alumni_year,
            "visible": "1",
            "description": (
                f"{alumni_year} alumni/archive cohort generated from final-grade cohort "
                f"{cohort.get('cohort_code', '')}."
            ),
        }
        generated.append([row.get(header, "") for header in headers])
    sheets["58_alumni_cohorts"] = (headers, generated)


def workbook_current_year(sheets: dict[str, tuple[list[str], list[list[str]]]], year: str) -> str:
    for row in rows_for(sheets, "02_academic_years"):
        if row.get("is_current") == "1" and row.get("academic_year"):
            return row["academic_year"]
    return year


def grade_band_policy_for(
    grade_code: str,
    stream_code: str,
    policies: list[dict[str, str]],
) -> dict[str, str]:
    for policy in policies:
        scope = policy.get("grade_codes", "")
        tokens = [token.strip() for token in scope.replace(",", "|").split("|") if token.strip()]
        if any(applies_token_matches(token, grade_code, stream_code) for token in tokens):
            return policy
    return {}


def auto_generate_attendance_policy(sheets: dict[str, tuple[list[str], list[list[str]]]], year: str) -> None:
    headers, rows = sheets.get("attendance_policy", ([], []))
    if not headers:
        return
    grades = rows_for(sheets, "05_grades")
    policies = rows_for(sheets, "34_grade_band_adjust")
    if not grades or not policies:
        return

    academic_year = workbook_current_year(sheets, year)
    generated: list[list[str]] = []
    for grade in grades:
        grade_code = grade.get("grade_code", "")
        if not grade_code:
            continue
        policy = grade_band_policy_for(grade_code, "", policies)
        row = {
            "academic_year": academic_year,
            "grade_code": grade_code,
            "minimum_attendance_percent": policy.get("minimum_attendance_percent", ""),
            "notes": policy.get("attendance_notes", ""),
        }
        generated.append([row.get(header, "") for header in headers])
    sheets["attendance_policy"] = (headers, generated)


def apply_generated_matrices(sheets: dict[str, tuple[list[str], list[list[str]]]], year: str) -> None:
    auto_generate_subject_matrix(sheets)
    auto_generate_alumni_cohorts(sheets, year)
    auto_generate_attendance_policy(sheets, year)


def convert(workbook: Path, output: Path, year: str, skip_example_row: bool = False) -> int:
    if not workbook.exists():
        raise FileNotFoundError(f"Workbook does not exist: {workbook}")
    wb = load_workbook(workbook, read_only=True, data_only=True)
    sheets: dict[str, tuple[list[str], list[list[str]]]] = {}
    for entry in SOURCE_FILES:
        sheet_name = entry["sheet"]
        if sheet_name not in wb.sheetnames:
            if entry.get("required"):
                raise RuntimeError(f"Missing required sheet: {sheet_name}")
            continue
        sheets[sheet_name] = sheet_to_rows(wb[sheet_name], skip_example_row=skip_example_row)

    apply_generated_matrices(sheets, year)

    targets: dict[Path, tuple[list[str], list[list[str]]]] = {}
    for entry in SOURCE_FILES:
        sheet_name = entry["sheet"]
        if sheet_name not in sheets:
            continue
        headers, rows = sheets[sheet_name]
        target = output / resolve_tokens(entry["source"], year)
        if target in targets:
            current_headers, current_rows = targets[target]
            if rows:
                if current_headers != headers:
                    raise RuntimeError(f"Duplicate target header mismatch for {target}")
                current_rows.extend(rows)
            continue
        targets[target] = (headers, rows)

    for target, (headers, rows) in targets.items():
        target.parent.mkdir(parents=True, exist_ok=True)
        with target.open("w", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows(rows)
    return len(targets)


def main():
    parser = argparse.ArgumentParser(description="Generate source CSV folder structure from the master import Excel workbook.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--year", default="2026-2027")
    parser.add_argument(
        "--skip-example-row",
        action="store_true",
        help="Skip metadata example_row values for pure template/review workbooks. Operational ODS workbooks should not use this.",
    )
    args = parser.parse_args()

    try:
        written = convert(Path(args.workbook), Path(args.output), args.year, skip_example_row=args.skip_example_row)
    except Exception as exc:
        print(f"Excel to CSV conversion failed: {exc}", file=sys.stderr)
        sys.exit(1)
    print(f"Generated {written} CSV files under {args.output}")


if __name__ == "__main__":
    main()
