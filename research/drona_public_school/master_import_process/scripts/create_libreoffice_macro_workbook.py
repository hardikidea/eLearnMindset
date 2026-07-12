#!/usr/bin/env python3
from __future__ import annotations

import argparse
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill


SOFFICE = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
MANIFEST_NS = "urn:oasis:names:tc:opendocument:xmlns:manifest:1.0"
MACRO_SOURCE = Path(__file__).with_name("libreoffice_master_tools.bas")
MAX_STATUS_ROWS = 100000


HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
CONTROL_FILL = PatternFill("solid", fgColor="DDEBF7")
AUTOMATIC_FILL = PatternFill("solid", fgColor="FFF2CC")
MANUAL_FILL = PatternFill("solid", fgColor="E7E6E6")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")


AUTOMATIC_SHEETS = {
    "13_grade_subject_matrix",
    "14_categories",
    "15_optional_year_category_model",
    "16_courses",
    "17_courses_with_templatecourse_",
    "18_cohorts",
    "19_groups",
    "26_cohort_members",
    "29_enrolments",
    "33_summary",
    "41_course_template_application",
    "50_academic_year_rollover_check",
    "55_student_academic_history_tem",
    "56_student_promotion_plan_2027_",
    "58_next_year_courses_2027_2028",
    "59_next_year_cohorts_2027_2028",
    "60_next_year_groups_2027_2028",
    "61_next_year_enrolments_2027_20",
    "62_alumni_cohorts_2027",
    "63_archive_policy",
    "65_compatibility_matrix",
    "66_assessment_plan",
    "67_attendance_policy",
    "68_course_certificates",
    "69_course_final_exams",
    "70_course_term_exams",
    "72_gradebook_weights",
}


AUTOMATIC_MACROS = {
    "13_grade_subject_matrix": "GenerateGradeSubjectMatrix",
    "14_categories": "GenerateCategories",
    "15_optional_year_category_model": "GenerateOptionalYearCategoryModel",
    "16_courses": "GenerateCourses",
    "17_courses_with_templatecourse_": "GenerateCoursesWithTemplateUpload",
    "18_cohorts": "GenerateCohorts",
    "19_groups": "GenerateGroups",
    "26_cohort_members": "GenerateCohortMembers",
    "29_enrolments": "GenerateEnrolments",
    "33_summary": "GenerateSummary",
    "41_course_template_application": "GenerateCourseTemplateApplication",
    "50_academic_year_rollover_check": "GenerateRolloverChecklist",
    "55_student_academic_history_tem": "GenerateStudentAcademicHistory",
    "56_student_promotion_plan_2027_": "GenerateStudentPromotionPlan",
    "58_next_year_courses_2027_2028": "GenerateNextYearCourses",
    "59_next_year_cohorts_2027_2028": "GenerateNextYearCohorts",
    "60_next_year_groups_2027_2028": "GenerateNextYearGroups",
    "61_next_year_enrolments_2027_20": "GenerateNextYearEnrolments",
    "62_alumni_cohorts_2027": "GenerateAlumniCohorts",
    "63_archive_policy": "GenerateArchivePolicy",
    "65_compatibility_matrix": "GenerateCompatibilityMatrix",
    "66_assessment_plan": "GenerateAssessmentPlan",
    "67_attendance_policy": "GenerateAttendancePolicy",
    "68_course_certificates": "GenerateCourseCertificates",
    "69_course_final_exams": "GenerateCourseFinalExams",
    "70_course_term_exams": "GenerateCourseTermExams",
    "72_gradebook_weights": "GenerateGradebookWeights",
}


ID_PATTERNS = {
    "13_grade_subject_matrix": "<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<SUBJECT_CODE>",
    "14_categories": "<TRUST_CODE>_<BOARD_CODE>_<SCHOOL_CODE>_<YYYY_YYYY>_<MEDIUM_CODE>_<GRADE_CODE>_<STREAM_CODE>",
    "15_optional_year_category_model": "<TRUST_CODE>_<BOARD_CODE>_<SCHOOL_CODE>_<YYYY_YYYY>_<MEDIUM_CODE>_<GRADE_CODE>_<STREAM_CODE>",
    "16_courses": "<SCHOOL_CODE>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<SUBJECT_CODE>-<START_YEAR>",
    "17_courses_with_templatecourse_": "<SCHOOL_CODE>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<SUBJECT_CODE>-<START_YEAR>",
    "18_cohorts": "<SCHOOL_CODE>-<START_YEAR>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<DIVISION_CODE>",
    "19_groups": "<COURSE_CODE>-<DIVISION_CODE>",
    "26_cohort_members": "<COHORT_CODE> + <STUDENT_USERNAME>",
    "29_enrolments": "<COURSE_CODE> + <COHORT_CODE> + cohort_sync",
    "33_summary": "Generated pack health summary from workbook source rows",
    "41_course_template_application": "<COURSE_CODE> + <TEMPLATECOURSE>",
    "50_academic_year_rollover_check": "Generated standard academic-year rollover checklist",
    "55_student_academic_history_tem": "<STUDENT_USERNAME> + <ACADEMIC_YEAR>",
    "56_student_promotion_plan_2027_": "<STUDENT_USERNAME> + <NEXT_ACADEMIC_YEAR>",
    "58_next_year_courses_2027_2028": "<SCHOOL_CODE>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<SUBJECT_CODE>-<NEXT_START_YEAR>",
    "59_next_year_cohorts_2027_2028": "<SCHOOL_CODE>-<NEXT_START_YEAR>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<DIVISION_CODE>",
    "60_next_year_groups_2027_2028": "<NEXT_COURSE_CODE>-<DIVISION_CODE>",
    "61_next_year_enrolments_2027_20": "<NEXT_COURSE_CODE> + <NEXT_COHORT_CODE> + cohort_sync",
    "62_alumni_cohorts_2027": "<SCHOOL_CODE>-ALUMNI-<NEXT_START_YEAR>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<DIVISION_CODE>",
    "63_archive_policy": "Generated standard archive and retention policy checklist",
    "65_compatibility_matrix": "Generated Moodle component compatibility matrix",
    "66_assessment_plan": "<COURSE_CODE> + assessment policy",
    "67_attendance_policy": "<GRADE_CODE> + <ACADEMIC_YEAR>",
    "68_course_certificates": "<COURSE_CODE> + certificate policy",
    "69_course_final_exams": "<COURSE_CODE> + FINAL",
    "70_course_term_exams": "<COURSE_CODE> + <TERM_CODE>",
    "72_gradebook_weights": "<COURSE_CODE> + gradebook weight policy",
}


def macro_hyperlink(macro_name: str, label: str) -> str:
    target = f"vnd.sun.star.script:Standard.MatrixTools.{macro_name}?language=Basic&location=document"
    return f'=HYPERLINK("{target}","{label}")'


def add_macro_guide_sheet(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    if "00_MACRO_GUIDE" in wb.sheetnames:
        del wb["00_MACRO_GUIDE"]
    ws = wb.create_sheet("00_MACRO_GUIDE", 1)
    ws.sheet_properties.tabColor = "F4B183"
    rows = [
        ["LibreOffice macro guide", "", ""],
        ["Macro module", "Standard.MatrixTools", ""],
        ["Main macro", "GenerateAllDerivedSheets", ""],
        ["Individual macros", "GenerateGradeSubjectMatrix, GenerateCategories, GenerateCourses, GenerateCoursesWithTemplateUpload, GenerateCohorts, GenerateGroups, GenerateCohortMembers, GenerateEnrolments, GenerateSummary, GenerateCourseTemplateApplication, GenerateRolloverChecklist, GenerateCourseCertificates, GenerateCourseFinalExams, GenerateCourseTermExams, GenerateGradebookWeights, GenerateOptionalYearCategoryModel, GenerateStudentAcademicHistory, GenerateStudentPromotionPlan, GenerateAssessmentPlan, GenerateAttendancePolicy, GenerateNextYearCourses, GenerateNextYearCohorts, GenerateNextYearGroups, GenerateNextYearEnrolments, GenerateAlumniCohorts, GenerateArchivePolicy, GenerateCompatibilityMatrix", ""],
        ["Derived target sheets", "13_grade_subject_matrix, 14_categories, 15_optional_year_category_model, 16_courses, 17_courses_with_templatecourse_, 18_cohorts, 19_groups, 26_cohort_members, 29_enrolments, 33_summary, 41_course_template_application, 50_academic_year_rollover_check, 55_student_academic_history_tem, 56_student_promotion_plan_2027_, 66_assessment_plan, 67_attendance_policy, 68_course_certificates, 69_course_final_exams, 70_course_term_exams, 72_gradebook_weights, 58_next_year_courses_2027_2028, 59_next_year_cohorts_2027_2028, 60_next_year_groups_2027_2028, 61_next_year_enrolments_2027_20, 62_alumni_cohorts_2027, 63_archive_policy, 65_compatibility_matrix", ""],
        ["How to run all", "Open this .ods in LibreOffice, enable macros when prompted, then use Tools > Macros > Run Macro > current document > Standard > MatrixTools > GenerateAllDerivedSheets.", ""],
        ["How to run one sheet", "Use the matching Generate* macro when you only changed one dependency area, for example GenerateCohorts after changing divisions or GenerateCourseCertificates after changing principal/course details.", ""],
        ["Formula behavior", "Macros create the required row structure and write formulas into derived cells. If a referenced master value changes later, existing derived rows recalculate automatically.", ""],
        ["Stream applies_to format", "Use pipe-delimited grade scopes in streams, for example STD01-STD10 or STD11_SCI|STD12_SCI. Comma-separated values are accepted only for older workbooks and should not be used for new data.", ""],
        ["Status sheet", "Use the status worksheet for data-health review. It shows manual versus automatic sheets, live row counts, PASSED/FAILED macro-derived checks, and action links for refresh, clear and reset/regenerate.", ""],
        ["When to rerun macros", "Rerun macros when row structure changes, for example a new academic year, grade, stream, medium, subject, division, course, or student registration row.", ""],
        ["Safety", "The macros clear and rebuild data rows only in derived sheets. Manual registration sheets for staff, students, parents, parent links and role assignments are intentionally not regenerated.", ""],
        ["Rule summary", "ID numbers follow the pack formulas: course_code = SCHOOL-BOARD-MEDIUM-GRADE-STREAM-SUBJECT-YEAR; cohort_code = SCHOOL-YEAR-BOARD-MEDIUM-GRADE-STREAM-DIVISION; group_idnumber = COURSE_CODE-DIVISION.", ""],
    ]
    for row in rows:
        ws.append(row)
    fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, size=14)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 120
    ws.column_dimensions["C"].width = 16
    wb.save(xlsx_path)


def excel_quote(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def source_sheet_rows(wb) -> list[dict[str, str]]:
    if "04_SOURCE_FILES" not in wb.sheetnames:
        return [
            {"source_csv": f"{sheet}.csv", "worksheet": sheet}
            for sheet in wb.sheetnames
            if sheet not in {"status", "00_MACRO_GUIDE"}
        ]
    ws = wb["04_SOURCE_FILES"]
    headers = {
        str(ws.cell(1, column).value or "").strip(): column
        for column in range(1, ws.max_column + 1)
    }
    rows: list[dict[str, str]] = []
    for row_index in range(2, ws.max_row + 1):
        worksheet = str(ws.cell(row_index, headers["worksheet"]).value or "").strip()
        source_csv = str(ws.cell(row_index, headers["source_csv"]).value or "").strip()
        if worksheet:
            rows.append({"source_csv": source_csv, "worksheet": worksheet})
    return rows


def status_expected_formula(sheet: str, row_by_sheet: dict[str, int]) -> str | None:
    def c(name: str) -> str:
        return f"C{row_by_sheet[name]}"

    formulas = {
        "13_grade_subject_matrix": lambda: c("16_courses"),
        "15_optional_year_category_model": lambda: c("14_categories"),
        "16_courses": lambda: c("13_grade_subject_matrix"),
        "17_courses_with_templatecourse_": lambda: c("16_courses"),
        "19_groups": lambda: f'{c("16_courses")}*COUNTA(\'11_divisions\'!A2:A{MAX_STATUS_ROWS})',
        "26_cohort_members": lambda: c("24_users_students"),
        "29_enrolments": lambda: c("19_groups"),
        "33_summary": lambda: "8",
        "41_course_template_application": lambda: c("16_courses"),
        "50_academic_year_rollover_check": lambda: "8",
        "55_student_academic_history_tem": lambda: c("24_users_students"),
        "56_student_promotion_plan_2027_": lambda: c("24_users_students"),
        "58_next_year_courses_2027_2028": lambda: c("16_courses"),
        "59_next_year_cohorts_2027_2028": lambda: c("18_cohorts"),
        "60_next_year_groups_2027_2028": lambda: c("19_groups"),
        "61_next_year_enrolments_2027_20": lambda: c("29_enrolments"),
        "62_alumni_cohorts_2027": lambda: f'COUNTIF(\'18_cohorts\'!H2:H{MAX_STATUS_ROWS},"STD12")',
        "63_archive_policy": lambda: "4",
        "65_compatibility_matrix": lambda: "18",
        "66_assessment_plan": lambda: c("16_courses"),
        "67_attendance_policy": lambda: f"COUNTA('09_grades'!A2:A{MAX_STATUS_ROWS})",
        "68_course_certificates": lambda: c("16_courses"),
        "69_course_final_exams": lambda: c("16_courses"),
        "70_course_term_exams": lambda: (
            f'{c("16_courses")}*COUNTIFS(\'71_exam_terms\'!A2:A{MAX_STATUS_ROWS},"<>FINAL",'
            f'\'71_exam_terms\'!A2:A{MAX_STATUS_ROWS},"<>")'
        ),
        "72_gradebook_weights": lambda: c("16_courses"),
    }
    factory = formulas.get(sheet)
    return factory() if factory else None


def add_status_sheet_to_workbook(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    if "status" in wb.sheetnames:
        del wb["status"]
    rows = source_sheet_rows(wb)
    ws = wb.create_sheet("status", 4)
    ws.sheet_properties.tabColor = "9DC3E6"
    ws.append(["type", "filename", "count", "status", "action"])
    control_rows = [
        (
            "automatic",
            "ALL_AUTOMATIC_SHEETS\nid_number_formula: rebuild every macro-generated sheet from current master rows",
            '=SUMIF(A7:A2000,"automatic",C7:C2000)',
            "CONTROL",
            macro_hyperlink("GenerateAllDerivedSheets", "Run all automatic macros"),
        ),
        (
            "automatic",
            "STATUS_ONLY\nid_number_formula: recalculate health formulas after manual review",
            "",
            "CONTROL",
            macro_hyperlink("RefreshStatus", "Refresh status"),
        ),
        (
            "automatic",
            "CLEAR_AUTOMATIC_SHEETS\nid_number_formula: remove data rows from macro-generated sheets",
            "",
            "CONTROL",
            macro_hyperlink("ClearAutomaticData", "Clear automatic data"),
        ),
        (
            "automatic",
            "RESET_AUTOMATIC_SHEETS\nid_number_formula: clear and regenerate all macro-generated sheets",
            "",
            "CONTROL",
            macro_hyperlink("ResetAutomaticData", "Reset and regenerate"),
        ),
        ("", "", "", "", ""),
    ]
    for row in control_rows:
        ws.append(row)

    row_by_sheet = {row["worksheet"]: index for index, row in enumerate(rows, start=7)}
    for row in rows:
        sheet = row["worksheet"]
        source_csv = row["source_csv"]
        row_index = row_by_sheet[sheet]
        row_type = "automatic" if sheet in AUTOMATIC_SHEETS else "manual"
        pattern = ID_PATTERNS.get(sheet, "Manual/source rows; follow worksheet column contract")
        filename = f"{source_csv}\nid_number_formula: {pattern}\nworksheet: {sheet}"
        count = f"=COUNTA({excel_quote(sheet)}!A2:A{MAX_STATUS_ROWS})"
        if row_type == "automatic":
            expected = status_expected_formula(sheet, row_by_sheet)
            if expected:
                status = f'=IF(C{row_index}=({expected}),"PASSED","FAILED")'
            else:
                status = f'=IF(C{row_index}>0,"PASSED","FAILED")'
            action = macro_hyperlink(AUTOMATIC_MACROS[sheet], "Run macro")
        else:
            status = "MANUAL"
            action = "Manual edit"
        ws.append([row_type, filename, count, status, action])

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    for column, width in {"A": 16, "B": 82, "C": 18, "D": 16, "E": 28}.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2, max_row=5):
        for cell in row:
            cell.fill = CONTROL_FILL
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        fill = AUTOMATIC_FILL if row[0].value == "automatic" else MANUAL_FILL
        for cell in row:
            cell.fill = fill
    ws.conditional_formatting.add(f"D7:D{ws.max_row}", FormulaRule(formula=['D7="PASSED"'], fill=PASS_FILL))
    ws.conditional_formatting.add(f"D7:D{ws.max_row}", FormulaRule(formula=['D7="FAILED"'], fill=FAIL_FILL))
    wb.save(xlsx_path)


def macro_source_code() -> str:
    if not MACRO_SOURCE.exists():
        raise FileNotFoundError(f"LibreOffice macro source not found: {MACRO_SOURCE}")
    return MACRO_SOURCE.read_text(encoding="utf-8")


def convert_xlsx_to_ods(source_xlsx: Path, output_ods: Path) -> Path:
    if not SOFFICE.exists():
        raise FileNotFoundError(f"LibreOffice soffice not found: {SOFFICE}")
    with tempfile.TemporaryDirectory(prefix="dps_macro_") as tmp:
        tmpdir = Path(tmp)
        temp_xlsx = tmpdir / source_xlsx.name
        shutil.copy2(source_xlsx, temp_xlsx)
        add_macro_guide_sheet(temp_xlsx)
        add_status_sheet_to_workbook(temp_xlsx)
        profile = tmpdir / "lo-profile"
        cmd = [
            str(SOFFICE),
            f"-env:UserInstallation=file://{profile}",
            "--headless",
            "--convert-to",
            "ods",
            "--outdir",
            str(tmpdir),
            str(temp_xlsx),
        ]
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        converted = tmpdir / f"{temp_xlsx.stem}.ods"
        if not converted.exists():
            raise FileNotFoundError(f"LibreOffice conversion did not create: {converted}")
        output_ods.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(converted, output_ods)
    return output_ods


def macro_files() -> dict[str, bytes]:
    macro_code = macro_source_code()
    module_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">\n'
        '<script:module xmlns:script="http://openoffice.org/2000/script" '
        'script:name="MatrixTools" script:language="StarBasic">'
        f"{escape(macro_code)}"
        "</script:module>"
    )
    script_lb = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="Standard" library:readonly="false" library:passwordprotected="false">
 <library:element library:name="MatrixTools"/>
</library:library>
"""
    script_lc = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:libraries PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "libraries.dtd">
<library:libraries xmlns:library="http://openoffice.org/2000/library" xmlns:xlink="http://www.w3.org/1999/xlink">
 <library:library library:name="Standard" xlink:href="Standard/script-lb.xml/" xlink:type="simple" library:link="false"/>
</library:libraries>
"""
    return {
        "Basic/Standard/MatrixTools.xba": module_xml.encode("utf-8"),
        "Basic/Standard/script-lb.xml": script_lb.encode("utf-8"),
        "Basic/script-lc.xml": script_lc.encode("utf-8"),
    }


def patch_manifest(manifest: bytes) -> bytes:
    ET.register_namespace("manifest", MANIFEST_NS)
    ET.register_namespace("loext", "urn:org:documentfoundation:names:experimental:office:xmlns:loext:1.0")
    root = ET.fromstring(manifest)
    full_path_attr = f"{{{MANIFEST_NS}}}full-path"
    media_type_attr = f"{{{MANIFEST_NS}}}media-type"
    existing = {entry.attrib.get(full_path_attr) for entry in root}
    entries = [
        ("Basic/", ""),
        ("Basic/Standard/", ""),
        ("Basic/script-lc.xml", "text/xml"),
        ("Basic/Standard/script-lb.xml", "text/xml"),
        ("Basic/Standard/MatrixTools.xba", "text/xml"),
    ]
    for path, media_type in entries:
        if path in existing:
            continue
        element = ET.SubElement(root, f"{{{MANIFEST_NS}}}file-entry")
        element.set(full_path_attr, path)
        element.set(media_type_attr, media_type)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def embed_macros(ods_path: Path) -> None:
    files = macro_files()
    temp_path = ods_path.with_suffix(".tmp.ods")
    with zipfile.ZipFile(ods_path, "r") as zin, zipfile.ZipFile(temp_path, "w") as zout:
        names = zin.namelist()
        mimetype = zin.read("mimetype") if "mimetype" in names else b"application/vnd.oasis.opendocument.spreadsheet"
        zout.writestr(zipfile.ZipInfo("mimetype"), mimetype, compress_type=zipfile.ZIP_STORED)
        for item in zin.infolist():
            if item.filename == "mimetype" or item.filename in files:
                continue
            data = zin.read(item.filename)
            if item.filename == "META-INF/manifest.xml":
                data = patch_manifest(data)
            zout.writestr(item, data)
        for path, data in files.items():
            zout.writestr(path, data, compress_type=zipfile.ZIP_DEFLATED)
    temp_path.replace(ods_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Create LibreOffice ODS workbook with MatrixTools macros.")
    parser.add_argument(
        "--source-xlsx",
        default="research/drona_public_school/build/master_excel/school_master_pack_2026_2027_full_predefined_master.xlsx",
    )
    parser.add_argument(
        "--output-ods",
        default="research/drona_public_school/build/master_excel/school_master_pack_2026_2027_full_predefined_master_macros.ods",
    )
    args = parser.parse_args()
    output = convert_xlsx_to_ods(Path(args.source_xlsx), Path(args.output_ods))
    embed_macros(output)
    print(f"Created LibreOffice macro workbook: {output}")


if __name__ == "__main__":
    main()
