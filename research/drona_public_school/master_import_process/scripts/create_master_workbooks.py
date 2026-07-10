#!/usr/bin/env python3
from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Border, Font, PatternFill, Alignment, Side
from openpyxl.utils import get_column_letter

from common import PACK_ROOT, PROCESS_ROOT, SOURCE_FILES, expected_headers, manifest_rows, read_rows, source_path


TEMPLATE_VERSION = "2026.07.07.3"
MAX_INPUT_ROWS = 2000
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
META_FILL = PatternFill("solid", fgColor="FFF2CC")
MANIFEST_FILL = PatternFill("solid", fgColor="D9EAD3")
EXAMPLE_FILL = PatternFill("solid", fgColor="EADCF8")
SUMMARY_FILL = PatternFill("solid", fgColor="EAF3F8")
PATTERN_FILL = PatternFill("solid", fgColor="F4F7EC")
MANDATORY_FILL = PatternFill("solid", fgColor="FCE4D6")
REQUIRED_IF_USED_FILL = PatternFill("solid", fgColor="FFF2CC")
OPTIONAL_FILL = PatternFill("solid", fgColor="E7E6E6")
SCHOOL_FILL = PatternFill("solid", fgColor="DDEBF7")
PARENT_FILL = PatternFill("solid", fgColor="EADCF8")
STUDENT_FILL = PatternFill("solid", fgColor="E2F0D9")
STAFF_FILL = PatternFill("solid", fgColor="D9EAD3")
ACADEMIC_FILL = PatternFill("solid", fgColor="FCE4D6")
SYSTEM_FILL = PatternFill("solid", fgColor="D9EAF7")
HELP_FILL = PatternFill("solid", fgColor="F3F6FA")
ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")
WARNING_FILL = PatternFill("solid", fgColor="FFEB9C")
HELPER_FILL = PatternFill("solid", fgColor="E8F3FF")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
CONTROL_FILL = PatternFill("solid", fgColor="DDEBF7")
SYSTEM_TAB = "9DC3E6"
REQUIRED_TAB = "F4B183"
OPTIONAL_TAB = "A6A6A6"
GENERATED_TAB = "A9D18E"
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)


LOOKUP_FIELDS = {
    "academic_year": "academic_year",
    "auth": "auth",
    "board_code": "board_code",
    "category_code": "category_code",
    "category_idnumber": "category_idnumber",
    "context_category_code": "category_code",
    "parent_category_code": "category_code",
    "cohort1": "cohort_code",
    "cohort_code": "cohort_code",
    "context_type": "context_type",
    "course_code": "course_code",
    "course_shortname": "course_shortname",
    "division_code": "division_code",
    "enablecompletion": "boolean",
    "enabled": "boolean",
    "enrolment_method": "enrolment_method",
    "format": "course_format",
    "grade_code": "grade_code",
    "group_idnumber": "group_idnumber",
    "groupmode": "groupmode",
    "groupmodeforce": "boolean",
    "is_compulsory": "boolean",
    "is_current": "boolean",
    "is_elective": "boolean",
    "medium_code": "medium_code",
    "moodle_role_shortname": "role_shortname",
    "parent_username": "parent_username",
    "relationship": "relationship",
    "role_shortname": "role_shortname",
    "school_code": "school_code",
    "showgrades": "boolean",
    "showreports": "boolean",
    "stream_code": "stream_code",
    "student_username": "student_username",
    "subject_code": "subject_code",
    "templatecourse": "templatecourse",
    "username": "username",
    "visible": "boolean",
}


STANDARD_PREFILL_SHEETS = {
    "02_academic_years",
    "03_boards",
    "04_mediums",
    "05_grades",
    "06_streams",
    "07_divisions",
    "08_subjects",
    "16_profile_fields",
    "17_custom_roles",
    "18_role_guidelines",
    "26_lookup_values",
    "27_validation_rules",
    "28_source_refs",
    "29_summary",
    "30_master_template",
    "31_template_sections",
    "32_template_activities",
    "33_template_gradebook",
    "34_grade_band_adjust",
    "35_subject_adjust",
    "36_completion_defaults",
    "38_template_custom_fields",
    "39_template_review",
    "40_certificate_policy",
    "41_report_access",
    "42_test_coverage",
    "44_transition_models",
    "45_promotion_rules",
    "46_rollover_checklist",
    "47_promotion_policy",
    "48_promotion_status",
    "49_promotion_validation",
    "50_student_status",
    "58_alumni_cohorts",
    "59_archive_policy",
    "60_improvement_backlog",
    "61_compatibility",
    "attendance_policy",
    "exam_terms",
}


REQUIREMENT_FILLS = {
    "Mandatory": MANDATORY_FILL,
    "Required if used": REQUIRED_IF_USED_FILL,
    "Optional": OPTIONAL_FILL,
}


AUDIENCE_FILLS = {
    "School-specific setup": SCHOOL_FILL,
    "Parent/family data": PARENT_FILL,
    "Student registration": STUDENT_FILL,
    "Staff registration": STAFF_FILL,
    "Academic/course setup": ACADEMIC_FILL,
    "Moodle/system reference": SYSTEM_FILL,
}


OPTIONAL_FIELD_NAMES = {
    "address",
    "address_line2",
    "advantages",
    "approved_by",
    "artifact_url",
    "attribution",
    "avoid",
    "description",
    "download_url",
    "example",
    "example_fix",
    "implementation_hint",
    "notes",
    "phone",
    "source_note",
    "source_url",
    "summary",
    "tags",
    "website",
}


GLOBAL_MANDATORY_FIELDS = {
    "academic_year",
    "auth",
    "board_code",
    "category_code",
    "category_type",
    "cohort_code",
    "course_code",
    "course_shortname",
    "division_code",
    "email",
    "firstname",
    "fullname",
    "grade_code",
    "group_idnumber",
    "group_name",
    "idnumber",
    "lastname",
    "medium_code",
    "name",
    "password",
    "role_shortname",
    "school_code",
    "shortname",
    "stream_code",
    "subject_code",
    "username",
}


MANDATORY_FIELDS_BY_SHEET = {
    "01_school_master": {
        "trust_code", "trust_name", "school_code", "school_name", "city", "district", "state",
        "pincode", "email", "principal_username", "academic_year",
    },
    "02_academic_years": {"academic_year", "start_date", "end_date", "is_current"},
    "03_boards": {"board_code", "board_name", "board_type", "country", "state"},
    "04_mediums": {"medium_code", "medium_name", "language_code"},
    "05_grades": {"grade_code", "grade_name", "display_order", "stage", "moodle_label"},
    "06_streams": {"stream_code", "stream_name", "applies_to"},
    "07_divisions": {"division_code", "division_name", "display_order"},
    "08_subjects": {"subject_code", "subject_name", "subject_category"},
    "09_subject_matrix": {
        "board_code", "medium_code", "grade_code", "stream_code", "subject_code",
        "subject_name", "subject_category", "is_compulsory", "display_order",
    },
    "10_categories": {"category_code", "category_type", "name", "idnumber", "visible"},
    "11_optional_categories": {"category_code", "category_type", "name", "idnumber", "visible"},
    "12_courses": {
        "course_code", "fullname", "shortname", "idnumber", "category_code", "board_code",
        "school_code", "medium_code", "grade_code", "stream_code", "subject_code",
        "academic_year", "format", "numsections", "visible",
    },
    "13_courses_upload": {"shortname", "fullname", "idnumber", "category_idnumber", "visible", "format", "numsections"},
    "14_cohorts": {
        "cohort_code", "name", "idnumber", "context_category_code", "board_code", "school_code",
        "medium_code", "grade_code", "stream_code", "division_code", "academic_year", "visible",
    },
    "15_groups": {"course_code", "course_shortname", "group_name", "group_idnumber", "division_code"},
    "19_users_staff": {"username", "password", "firstname", "lastname", "email", "auth", "idnumber"},
    "20_users_students": {"username", "password", "firstname", "lastname", "email", "auth", "idnumber", "cohort1"},
    "21_users_parents": {"username", "password", "firstname", "lastname", "email", "auth", "idnumber"},
    "22_cohort_members": {"cohort_code", "username"},
    "23_role_assignments": {"username", "role_shortname", "context_type", "context_identifier"},
    "24_parent_links": {"parent_username", "student_username", "relationship"},
    "25_enrolments": {"course_code", "cohort_code", "role_shortname", "enrolment_method"},
}


SUMMARY_OVERRIDES = {
    "academic_year": "Academic year scope used for categories, courses, cohorts and enrolments.",
    "address": "Primary address text for the user or school record.",
    "address_line1": "First line of the school address.",
    "address_line2": "Second line of the school address, if required.",
    "affiliation_no": "Board affiliation or recognition number.",
    "auth": "Moodle authentication method for the user.",
    "board_code": "Short board code used across matrix, categories, courses and enrolments.",
    "board_name": "Display name of the education board.",
    "category_code": "Stable source category key used by scripts before Moodle category creation.",
    "category_idnumber": "Moodle category idnumber used for course upload references.",
    "category_path": "Human-readable Moodle category path.",
    "category_type": "Category level such as trust, school, academic year, medium, grade or stream.",
    "city": "City for the school or user profile.",
    "cohort_code": "Stable source cohort key for one class, stream and division.",
    "cohort1": "Cohort idnumber assigned to the user during upload.",
    "context_category_code": "Category code where the cohort should be created.",
    "course_code": "Stable source course key for a subject in a specific year.",
    "course_shortname": "Moodle course short name reference.",
    "department": "School department, grade band or staff department.",
    "description": "Short operator-facing description.",
    "display_order": "Numeric sorting order in generated data or UI lists.",
    "district": "District for school or user profile reporting.",
    "division_code": "Division/section code such as A, B or C.",
    "email": "Unique email address used by Moodle user accounts.",
    "enablecompletion": "Whether Moodle course completion tracking is enabled.",
    "firstname": "User first name.",
    "format": "Moodle course format such as topics.",
    "fullname": "Moodle full course name.",
    "grade_code": "Standard/grade code used across course, cohort and group records.",
    "group_idnumber": "Moodle group idnumber for course-specific division grouping.",
    "group_name": "Moodle group display name.",
    "groupmode": "Moodle group mode value for the course.",
    "groupmodeforce": "Whether Moodle forces the configured group mode.",
    "idnumber": "Moodle idnumber or source idnumber depending on the sheet context.",
    "institution": "School or trust name stored in user profile.",
    "is_compulsory": "Marks whether a subject is compulsory in the matrix.",
    "is_current": "Marks the active academic year.",
    "is_elective": "Marks whether a subject is elective in the matrix.",
    "lang": "Preferred Moodle language code.",
    "language_code": "ISO-style language code for a teaching medium.",
    "lastname": "User last name.",
    "medium_code": "Teaching medium code such as GUJ, ENG or HIN.",
    "moodle_role_shortname": "Moodle role shortname used in role assignments.",
    "name": "Display name for the target record.",
    "notes": "Short notes for operators and reviewers.",
    "numsections": "Number of sections created in the course.",
    "parent_category_code": "Parent category code; blank only for the root category.",
    "parent_username": "Parent username linked to a student.",
    "password": "Initial Moodle password when creating users.",
    "phone": "Contact phone number.",
    "pincode": "Postal PIN code.",
    "principal_username": "Username of the principal user.",
    "role_shortname": "Role shortname assigned in the target context.",
    "school_code": "Short school code used as the main identity token.",
    "school_name": "School display name.",
    "school_type": "School type such as private, trust-run or government aided.",
    "shortname": "Moodle short name, usually compact and unique.",
    "source_note": "Source note explaining how the row was selected.",
    "source_url": "Reference URL for the source record.",
    "stage": "School stage such as primary, secondary or higher secondary.",
    "state": "State for school or user profile reporting.",
    "stream_code": "Stream code such as GEN, SCI or COM.",
    "subject_category": "Subject category such as core, language, practical or elective.",
    "subject_code": "Subject code used across course and matrix records.",
    "subject_name": "Subject display name.",
    "summary": "Moodle course summary or operator-facing detail.",
    "tags": "Comma-separated course tags.",
    "templatecourse": "Moodle template course shortname used for course creation.",
    "trust_code": "Trust code used in root category and naming formulas.",
    "trust_name": "Trust display name.",
    "udise_code": "UDISE or school registration code.",
    "username": "Unique Moodle username.",
    "visible": "Visibility flag used by Moodle and generated records.",
    "website": "School website URL.",
}


BOOLEAN_FIELDS = {
    "allow_activity_report_view",
    "allow_grade_view",
    "apply_completion_defaults",
    "apply_gradebook",
    "apply_sections",
    "certificate_email_students",
    "certificate_enabled",
    "certificate_verification_enabled",
    "compatible_moodle_5_0_2",
    "compatible_moodle_5_2_branch_502",
    "completion_relevant",
    "completion_reports",
    "completion_required",
    "default_visibility",
    "enablecompletion",
    "enabled",
    "expiry_enabled",
    "forceunique",
    "groupmodeforce",
    "is_compulsory",
    "is_current",
    "is_elective",
    "locked",
    "notify_parent",
    "required",
    "showgrades",
    "showreports",
    "visible",
}


INTEGER_FIELDS = {
    "default_extra_sections",
    "default_points",
    "display_order",
    "item_order",
    "minimum_attendance_percent",
    "numsections",
    "passing_grade_percent",
    "section_number",
    "sortorder",
}


BASE_PATTERNS = {
    "academic_year": "<START_YEAR>-<END_YEAR>, e.g. 2026-2027",
    "auth": "manual",
    "board_code": "<BOARD_CODE>, e.g. GSEB",
    "category_code": "<TRUST_CODE>_<BOARD_CODE>_<SCHOOL_CODE>_<YYYY_YYYY>_<MEDIUM_CODE>_<GRADE_CODE>_<STREAM_CODE>",
    "category_idnumber": "Must match 10_categories.idnumber",
    "category_path": "<Trust> / <Board> / <School> / <Academic Year> / <Medium> / <Grade> / <Stream>",
    "cohort1": "Must match 14_cohorts.idnumber",
    "cohort_code": "<SCHOOL_CODE>-<START_YEAR>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<DIVISION_CODE>",
    "context_category_code": "Must match 10_categories.category_code",
    "course_code": "<SCHOOL_CODE>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<SUBJECT_CODE>-<START_YEAR>",
    "course_shortname": "Must match 12_courses.shortname",
    "division_code": "<DIVISION_CODE>, e.g. A",
    "email": "<unique-account>@<school-domain>",
    "end_date": "YYYY-MM-DD",
    "firstname": "Given name",
    "format": "topics",
    "grade_code": "STD<2_DIGITS>, e.g. STD05",
    "group_idnumber": "<COURSE_IDNUMBER>-<DIVISION_CODE>",
    "group_name": "<Division name> or <Course> - Division <DIVISION_CODE>",
    "idnumber": "<IDNUMBER_FORMULA_BY_SHEET>",
    "lang": "en",
    "language_code": "ISO language code, e.g. gu/en/hi",
    "lastname": "Family name",
    "medium_code": "<MEDIUM_CODE>, e.g. GUJ/ENG/HIN",
    "moodle_role_shortname": "student/teacher/editingteacher/principal/trustee_manager",
    "parent_category_code": "Must match 10_categories.category_code or blank for root",
    "parent_username": "Must match 21_users_parents.username",
    "password": "Initial password meeting Moodle complexity rules",
    "phone": "Indian phone number, digits only preferred",
    "pincode": "6 digit Indian PIN code",
    "role_shortname": "student/teacher/editingteacher/manager/custom role",
    "school_code": "<SCHOOL_CODE>, e.g. DPS",
    "shortname": "<SCHOOL_CODE>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<SUBJECT_CODE>-<YY>",
    "start_date": "YYYY-MM-DD",
    "stream_code": "<STREAM_CODE>, e.g. GEN/SCI/COM",
    "subject_code": "<SUBJECT_CODE>, e.g. MATH",
    "templatecourse": "Existing Moodle template course shortname",
    "trust_code": "<TRUST_CODE>",
    "username": "<school_code_lower>.<role>.<sequence_or_subject>",
    "visible": "1 visible, 0 hidden",
}


SHEET_IDNUMBER_PATTERNS = {
    "10_categories": "<TRUST_CODE>_<BOARD_CODE>_<SCHOOL_CODE>_<YYYY_YYYY>_<MEDIUM_CODE>_<GRADE_CODE>_<STREAM_CODE>",
    "11_optional_categories": "<TRUST_CODE>_<BOARD_CODE>_<SCHOOL_CODE>_<YYYY_YYYY>_<MEDIUM_CODE>_<GRADE_CODE>_<STREAM_CODE>",
    "12_courses": "<SCHOOL_CODE>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<SUBJECT_CODE>-<START_YEAR>",
    "13_courses_upload": "<SCHOOL_CODE>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<SUBJECT_CODE>-<START_YEAR>",
    "14_cohorts": "<SCHOOL_CODE>-<START_YEAR>-<BOARD_CODE>-<MEDIUM_CODE>-<GRADE_CODE>-<STREAM_CODE>-<DIVISION_CODE>",
    "19_users_staff": "<SCHOOL_CODE>-TCH-<3_DIGIT_SEQUENCE> or staff id",
    "20_users_students": "<SCHOOL_CODE><YY>-<5_DIGIT_SEQUENCE>",
    "21_users_parents": "<SCHOOL_CODE>-PAR-<5_DIGIT_SEQUENCE>",
}


AUTOMATIC_SHEETS = {
    "09_subject_matrix",
    "10_categories",
    "11_optional_categories",
    "12_courses",
    "13_courses_upload",
    "14_cohorts",
    "15_groups",
    "22_cohort_members",
    "25_enrolments",
    "37_template_application",
    "51_academic_history",
    "52_promotion_plan",
    "54_next_year_courses",
    "55_next_year_cohorts",
    "56_next_year_groups",
    "57_next_year_enrolments",
    "assessment_plan",
    "attendance_policy",
    "course_certificates",
    "course_final_exams",
    "course_term_exams",
    "gradebook_weights",
}


AUTOMATIC_MACROS = {
    "09_subject_matrix": "GenerateGradeSubjectMatrix",
    "10_categories": "GenerateCategories",
    "11_optional_categories": "GenerateOptionalYearCategoryModel",
    "12_courses": "GenerateCourses",
    "13_courses_upload": "GenerateCoursesWithTemplateUpload",
    "14_cohorts": "GenerateCohorts",
    "15_groups": "GenerateGroups",
    "22_cohort_members": "GenerateCohortMembers",
    "25_enrolments": "GenerateEnrolments",
    "37_template_application": "GenerateCourseTemplateApplication",
    "51_academic_history": "GenerateStudentAcademicHistory",
    "52_promotion_plan": "GenerateStudentPromotionPlan",
    "54_next_year_courses": "GenerateNextYearCourses",
    "55_next_year_cohorts": "GenerateNextYearCohorts",
    "56_next_year_groups": "GenerateNextYearGroups",
    "57_next_year_enrolments": "GenerateNextYearEnrolments",
    "assessment_plan": "GenerateAssessmentPlan",
    "attendance_policy": "GenerateAttendancePolicy",
    "course_certificates": "GenerateCourseCertificates",
    "course_final_exams": "GenerateCourseFinalExams",
    "course_term_exams": "GenerateCourseTermExams",
    "gradebook_weights": "GenerateGradebookWeights",
}


def macro_hyperlink(macro_name: str, label: str) -> str:
    target = f"vnd.sun.star.script:Standard.MatrixTools.{macro_name}?language=Basic&location=document"
    return f'=HYPERLINK("{target}","{label}")'


def fit_columns(ws, max_width=52):
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 12
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def excel_quote(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def safe_range_name(name: str) -> str:
    safe = "".join(ch if ch.isalnum() else "_" for ch in name)
    if not safe or safe[0].isdigit():
        safe = f"lookup_{safe}"
    return safe[:200]


def unique_values(rows: list[dict[str, str]], field: str) -> list[str]:
    seen: set[str] = set()
    values: list[str] = []
    for row in rows:
        value = str(row.get(field, "")).strip()
        if value and value not in seen:
            seen.add(value)
            values.append(value)
    return values


def source_values(source_root: Path, year: str, source: str, field: str) -> list[str]:
    path = source_root / source.format(year=year, next_year=year_next(year))
    return unique_values(read_rows(path), field)


def build_lookup_values(year: str, source_root: Path) -> dict[str, list[str]]:
    values = {
        "auth": ["manual"],
        "boolean": ["1", "0"],
        "context_type": ["system", "category", "course"],
        "course_format": ["topics"],
        "enrolment_method": ["cohort_sync"],
        "groupmode": ["0", "1", "2"],
        "relationship": ["Guardian", "Father", "Mother", "Parent"],
    }
    values["academic_year"] = source_values(source_root, year, "master/academic_years.csv", "academic_year")
    values["board_code"] = source_values(source_root, year, "master/boards.csv", "board_code")
    values["medium_code"] = source_values(source_root, year, "master/mediums.csv", "medium_code")
    values["grade_code"] = source_values(source_root, year, "master/grades.csv", "grade_code")
    values["stream_code"] = source_values(source_root, year, "master/streams.csv", "stream_code")
    values["division_code"] = source_values(source_root, year, "master/divisions.csv", "division_code")
    values["subject_code"] = source_values(source_root, year, "master/subjects.csv", "subject_code")
    values["role_shortname"] = source_values(source_root, year, "master/roles.csv", "role_shortname") + [
        "student", "teacher", "editingteacher", "manager", "parent", "principal", "trustee_manager"
    ]
    school_rows = read_rows(source_root / "master" / "school.csv")
    values["school_code"] = unique_values(school_rows, "school_code")
    values["category_code"] = source_values(source_root, year, "years/{year}/categories.csv", "category_code")
    values["category_idnumber"] = source_values(source_root, year, "years/{year}/categories.csv", "idnumber")
    values["course_code"] = source_values(source_root, year, "years/{year}/courses.csv", "course_code")
    values["course_shortname"] = source_values(source_root, year, "years/{year}/courses.csv", "shortname")
    values["templatecourse"] = values["course_shortname"]
    values["cohort_code"] = source_values(source_root, year, "years/{year}/cohorts.csv", "cohort_code")
    values["group_idnumber"] = source_values(source_root, year, "years/{year}/groups.csv", "group_idnumber")
    values["student_username"] = source_values(source_root, year, "registration/combined/20_users_students.csv", "username")
    values["parent_username"] = source_values(source_root, year, "registration/combined/21_users_parents.csv", "username")
    staff = source_values(source_root, year, "registration/combined/19_users_staff.csv", "username")
    values["username"] = values["student_username"] + values["parent_username"] + staff
    return {key: sorted(dict.fromkeys(items)) for key, items in values.items() if items}


def add_defined_lookup_range(wb: Workbook, name: str, sheet_name: str, column: int, count: int) -> str:
    range_name = safe_range_name(f"lookup_{name}")
    last_row = max(2, count + 1)
    column_letter = get_column_letter(column)
    attr_text = f"{excel_quote(sheet_name)}!${column_letter}$2:${column_letter}${last_row}"
    wb.defined_names.add(DefinedName(range_name, attr_text=attr_text))
    return range_name


def add_lookup_sheet(wb: Workbook, lookup_values: dict[str, list[str]]) -> dict[str, str]:
    ws = wb.create_sheet("_lookups")
    ws.sheet_state = "hidden"
    ws.sheet_properties.tabColor = SYSTEM_TAB
    lookup_names: dict[str, str] = {}
    for column, (name, values) in enumerate(sorted(lookup_values.items()), start=1):
        ws.cell(row=1, column=column, value=name)
        for row_number, value in enumerate(values, start=2):
            ws.cell(row=row_number, column=column, value=value)
        lookup_names[name] = add_defined_lookup_range(wb, name, ws.title, column, len(values))
    fit_columns(ws)
    return lookup_names


def add_data_validation(ws, cell_range: str, range_name: str, allow_blank: bool) -> None:
    dv = DataValidation(type="list", formula1=f"={range_name}", allow_blank=allow_blank)
    dv.error = "Choose a value from the dropdown list or update the master sheet first."
    dv.errorTitle = "Invalid value"
    dv.prompt = "Choose a configured value."
    dv.promptTitle = "Configured values"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def row_has_data_formula(first_col: str, last_col: str, row: int) -> str:
    return f"COUNTA(${first_col}{row}:${last_col}{row})>0"


def apply_conditional_formatting(
    ws,
    headers: list[str],
    column_guides: list[tuple[str, str]],
    data_start_row: int,
) -> None:
    if not headers:
        return
    last_col = get_column_letter(len(headers))
    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        cell = f"{letter}{data_start_row}"
        requirement = column_guides[index - 1][0]
        target_range = f"{letter}{data_start_row}:{letter}{MAX_INPUT_ROWS}"
        if requirement in {"Mandatory", "Required if used"}:
            formula = f"AND({row_has_data_formula('A', last_col, data_start_row)},LEN(TRIM({cell}))=0)"
            ws.conditional_formatting.add(target_range, FormulaRule(formula=[formula], fill=ERROR_FILL))
        if header == "email":
            formula = f'AND(LEN(TRIM({cell}))>0,ISERROR(SEARCH("@",{cell})))'
            ws.conditional_formatting.add(target_range, FormulaRule(formula=[formula], fill=WARNING_FILL))
        if header in {"academic_year"}:
            formula = f'AND(LEN(TRIM({cell}))>0,NOT(ISNUMBER(SEARCH("-",{cell}))))'
            ws.conditional_formatting.add(target_range, FormulaRule(formula=[formula], fill=WARNING_FILL))


def header_index(headers: list[str], name: str) -> int | None:
    try:
        return headers.index(name) + 1
    except ValueError:
        return None


def cell_ref(headers: list[str], name: str, row: int) -> str | None:
    index = header_index(headers, name)
    if not index:
        return None
    return f"{get_column_letter(index)}{row}"


def join_formula(parts: list[str | None]) -> str | None:
    if any(part is None for part in parts):
        return None
    return "=" + '&"-"&'.join(parts)  # type: ignore[arg-type]


def add_formula_helpers(
    ws,
    sheet: str,
    headers: list[str],
    column_guides: list[tuple[str, str]],
    data_start_row: int,
) -> None:
    helper_start = max(len(headers), 18) + 2
    helper_columns: list[tuple[str, str]] = []
    last_data_col = get_column_letter(len(headers))
    required_refs = [
        f'{get_column_letter(index)}{data_start_row}=""'
        for index, guide in enumerate(column_guides, start=1)
        if guide[0] in {"Mandatory", "Required if used"}
    ]
    if required_refs:
        helper_columns.append((
            "helper_row_status",
            f'=IF(COUNTA($A{data_start_row}:${last_data_col}{data_start_row})=0,"",IF(OR({",".join(required_refs)}),"Missing required value","Ready"))',
        ))
    if sheet == "12_courses":
        course_parts = [
            cell_ref(headers, "school_code", data_start_row),
            cell_ref(headers, "board_code", data_start_row),
            cell_ref(headers, "medium_code", data_start_row),
            cell_ref(headers, "grade_code", data_start_row),
            cell_ref(headers, "stream_code", data_start_row),
            cell_ref(headers, "subject_code", data_start_row),
            f'LEFT({cell_ref(headers, "academic_year", data_start_row)},4)' if cell_ref(headers, "academic_year", data_start_row) else None,
        ]
        short_parts = course_parts[:-1] + [
            f'RIGHT(LEFT({cell_ref(headers, "academic_year", data_start_row)},4),2)' if cell_ref(headers, "academic_year", data_start_row) else None
        ]
        formula = join_formula(course_parts)
        short_formula = join_formula(short_parts)
        if formula:
            helper_columns.append(("helper_expected_course_code", formula))
        if short_formula:
            helper_columns.append(("helper_expected_shortname", short_formula))
    elif sheet == "14_cohorts":
        cohort_parts = [
            cell_ref(headers, "school_code", data_start_row),
            f'LEFT({cell_ref(headers, "academic_year", data_start_row)},4)' if cell_ref(headers, "academic_year", data_start_row) else None,
            cell_ref(headers, "board_code", data_start_row),
            cell_ref(headers, "medium_code", data_start_row),
            cell_ref(headers, "grade_code", data_start_row),
            cell_ref(headers, "stream_code", data_start_row),
            cell_ref(headers, "division_code", data_start_row),
        ]
        formula = join_formula(cohort_parts)
        if formula:
            helper_columns.append(("helper_expected_cohort_code", formula))
    elif sheet == "15_groups":
        course = cell_ref(headers, "course_code", data_start_row)
        division = cell_ref(headers, "division_code", data_start_row)
        if course and division:
            helper_columns.append(("helper_expected_group_idnumber", f'={course}&"-"&{division}'))
    elif sheet == "21_users_parents":
        username = cell_ref(headers, "username", data_start_row)
        if username:
            helper_columns.append(("helper_expected_parent_idnumber", f'=UPPER(LEFT({username},FIND(".",{username})-1))&"-PAR-"&RIGHT({username},5)'))

    for offset, (label, formula) in enumerate(helper_columns):
        column = helper_start + offset
        letter = get_column_letter(column)
        ws.cell(row=1, column=column, value="helper_columns_not_exported")
        ws.cell(row=2, column=column, value="Helper | Moodle/system reference")
        ws.cell(row=3, column=column, value=f"{label} is generated for operator review only.")
        ws.cell(row=4, column=column, value="Not exported to CSV")
        ws.cell(row=5, column=column, value="")
        for row in range(data_start_row, MAX_INPUT_ROWS + 1):
            ws.cell(row=row, column=column, value=formula.replace(str(data_start_row), str(row)))
        for row in range(1, min(MAX_INPUT_ROWS, data_start_row + 25) + 1):
            cell = ws.cell(row=row, column=column)
            cell.fill = HELPER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[letter].width = 28


def humanize_header(header: str) -> str:
    return header.replace("_", " ").strip()


def column_summary(header: str) -> str:
    if header in SUMMARY_OVERRIDES:
        return SUMMARY_OVERRIDES[header]
    label = humanize_header(header)
    if header.endswith("_code") or header == "code":
        return f"Stable code used to identify or reference {label}."
    if header.endswith("_idnumber") or header == "idnumber":
        return f"Moodle idnumber or external reference for {label}."
    if header.endswith("_username") or header == "username":
        return f"Moodle username reference for {label}."
    if header.endswith("_name") or header == "name":
        return f"Display name for {label}."
    if header.endswith("_date") or header in {"created_on", "effective_from", "effective_to"}:
        return f"Date value for {label}."
    if header in BOOLEAN_FIELDS or header.startswith("is_") or header.startswith("allow_") or header.startswith("apply_"):
        return f"Yes/no flag for {label}."
    if "percent" in header or header.endswith("_weight"):
        return f"Numeric percentage or weight for {label}."
    if header.endswith("_url"):
        return f"URL reference for {label}."
    if header.endswith("_type") or header.endswith("_mode"):
        return f"Controlled value for {label}."
    return f"Value for {label}."


def column_pattern(sheet: str, header: str) -> str:
    if header == "idnumber" and sheet in SHEET_IDNUMBER_PATTERNS:
        return SHEET_IDNUMBER_PATTERNS[sheet]
    if header in BASE_PATTERNS:
        return BASE_PATTERNS[header]
    label = humanize_header(header)
    if header.endswith("_code") or header == "code":
        return f"<{header.upper()}>; keep stable and reference from the matching master sheet"
    if header.endswith("_idnumber"):
        return f"Must match the referenced {label} record"
    if header.endswith("_username"):
        return f"Must match the referenced Moodle {label}"
    if header.endswith("_date") or header in {"created_on", "effective_from", "effective_to"}:
        return "YYYY-MM-DD"
    if header in BOOLEAN_FIELDS or header.startswith("is_") or header.startswith("allow_") or header.startswith("apply_"):
        return "1 yes, 0 no"
    if header in INTEGER_FIELDS:
        return "Whole number"
    if "percent" in header:
        return "0-100"
    if header.endswith("_weight"):
        return "Numeric weight; related weights should total 100 where applicable"
    if header.endswith("_url") or header in {"website", "download_url", "artifact_url"}:
        return "https://..."
    if header in {"email"}:
        return "Valid unique email address"
    return "Free text or controlled value; follow related reference columns when present"


def column_requirement(sheet: str, header: str, entry_required: bool) -> str:
    mandatory_fields = MANDATORY_FIELDS_BY_SHEET.get(sheet, set())
    if header in mandatory_fields or header in GLOBAL_MANDATORY_FIELDS:
        return "Mandatory" if entry_required else "Required if used"
    if header in OPTIONAL_FIELD_NAMES or header.startswith("optional_"):
        return "Optional"
    if header in BOOLEAN_FIELDS or header in INTEGER_FIELDS:
        return "Optional"
    return "Required if used" if not entry_required else "Optional"


def column_audience(sheet: str, header: str) -> str:
    if sheet in {"21_users_parents", "24_parent_links"} or "parent" in header or "guardian" in header:
        return "Parent/family data"
    if sheet == "20_users_students" or "student" in header or header in {"cohort1", "new_roll_no"}:
        return "Student registration"
    if sheet == "19_users_staff" or any(token in header for token in ("teacher", "principal", "trustee", "staff")):
        return "Staff registration"
    if sheet in {
        "01_school_master", "02_academic_years", "03_boards", "04_mediums", "05_grades",
        "06_streams", "07_divisions", "08_subjects", "16_profile_fields", "17_custom_roles",
        "18_role_guidelines", "26_lookup_values", "27_validation_rules", "28_source_refs",
        "29_summary",
    }:
        return "School-specific setup"
    if header in {
        "trust_code", "school_code", "board_code", "medium_code", "grade_code", "stream_code",
        "division_code", "subject_code", "academic_year",
    }:
        return "School-specific setup"
    if sheet.startswith(("30_", "31_", "32_", "33_", "34_", "35_", "36_", "37_", "38_", "39_", "40_", "41_", "42_", "43_")):
        return "Academic/course setup"
    if sheet in {
        "09_subject_matrix", "10_categories", "11_optional_categories", "12_courses",
        "13_courses_upload", "14_cohorts", "15_groups", "22_cohort_members",
        "23_role_assignments", "25_enrolments", "assessment_plan", "attendance_policy",
        "course_certificates", "course_final_exams", "course_term_exams", "exam_terms",
        "gradebook_weights",
    }:
        return "Academic/course setup"
    return "Moodle/system reference"


def column_guide(sheet: str, header: str, entry_required: bool) -> tuple[str, str]:
    requirement = column_requirement(sheet, header, entry_required)
    audience = column_audience(sheet, header)
    return requirement, audience


def style_data_sheet(
    ws,
    guide_row: int,
    summary_row: int,
    pattern_row: int,
    header_row: int,
    example_row: int | None,
    column_guides: list[tuple[str, str]],
):
    ws.freeze_panes = f"A{header_row + 1}"
    for cell in ws[1]:
        cell.fill = META_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = THIN_BORDER
    for index, cell in enumerate(ws[guide_row], start=1):
        requirement, audience = column_guides[index - 1] if index <= len(column_guides) else ("Optional", "Moodle/system reference")
        cell.fill = AUDIENCE_FILLS[audience]
        cell.font = Font(bold=True, color="1F2937")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
    for cell in ws[summary_row]:
        cell.fill = SUMMARY_FILL
        cell.font = Font(italic=True, color="4A5568")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
    for cell in ws[pattern_row]:
        cell.fill = PATTERN_FILL
        cell.font = Font(italic=True, color="385723")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
    for index, cell in enumerate(ws[header_row], start=1):
        requirement, _audience = column_guides[index - 1] if index <= len(column_guides) else ("Optional", "Moodle/system reference")
        cell.fill = REQUIREMENT_FILLS[requirement]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = THIN_BORDER
    if example_row and ws.max_row >= example_row:
        for cell in ws[example_row]:
            cell.fill = EXAMPLE_FILL
            cell.font = Font(italic=True)
            cell.alignment = Alignment(wrap_text=True)
            cell.border = THIN_BORDER
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = THIN_BORDER
    fit_columns(ws)


def add_manifest(wb, year):
    ws = wb.create_sheet("_manifest", 0)
    ws.sheet_properties.tabColor = SYSTEM_TAB
    headers = ["sheet_name", "source_csv", "ordered_csv", "required", "purpose"]
    ws.append(headers)
    for row in manifest_rows(year):
        ws.append([row[h] for h in headers])
    for cell in ws[1]:
        cell.fill = MANIFEST_FILL
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    fit_columns(ws)


def add_dashboard(wb, year):
    ws = wb.create_sheet("_dashboard", 0)
    ws.sheet_properties.tabColor = SYSTEM_TAB
    rows = [
        ("Master import dashboard", "", "", ""),
        ("Academic year", year, "", ""),
        ("Template version", TEMPLATE_VERSION, "", ""),
        ("Metric", "Formula", "Purpose", "Action"),
        ("Students", "=COUNTA('20_users_students'!A7:A2000)", "Student accounts to import.", "Review student registration rows."),
        ("Parents", "=COUNTA('21_users_parents'!A7:A2000)", "Parent accounts to import.", "Review parent registration rows."),
        ("Staff", "=COUNTA('19_users_staff'!A7:A2000)", "Staff accounts to import.", "Review teacher/principal/trustee rows."),
        ("Courses", "=COUNTA('12_courses'!A7:A2000)", "Courses to create/update.", "Review generated course codes."),
        ("Cohorts", "=COUNTA('14_cohorts'!A7:A2000)", "Class/division cohorts.", "Review student grouping."),
        ("Groups", "=COUNTA('15_groups'!D7:D2000)", "Course division groups.", "Review group idnumbers."),
        ("Parent links", "=COUNTA('24_parent_links'!A7:A2000)", "Parent-to-student relationships.", "Confirm every active student has expected parent links."),
        ("Enrolments", "=COUNTA('25_enrolments'!A7:A2000)", "Course/cohort enrolment rows.", "Confirm all courses have required cohort sync rows."),
        ("Certificates", "=COUNTA('course_certificates'!A7:A2000)", "Course certificate configuration.", "Confirm certificate plugin is available before live import."),
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].fill = HELP_FILL
    for cell in ws[4]:
        cell.fill = MANIFEST_FILL
        cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A5"
    fit_columns(ws, max_width=72)


def status_pattern(entry: dict) -> str:
    sheet = entry["sheet"]
    if sheet in SHEET_IDNUMBER_PATTERNS:
        return SHEET_IDNUMBER_PATTERNS[sheet]
    source = entry["source"]
    if sheet in AUTOMATIC_SHEETS:
        return f"Generated from {source}; rerun macro when dependency row counts change"
    return f"Manual/source data from {source}"


def status_expected_formula(sheet: str, row_by_sheet: dict[str, int]) -> str | None:
    def c(name: str) -> str:
        return f"C{row_by_sheet[name]}"

    formulas = {
        "09_subject_matrix": lambda: c("12_courses"),
        "11_optional_categories": lambda: c("10_categories"),
        "12_courses": lambda: c("09_subject_matrix"),
        "13_courses_upload": lambda: c("12_courses"),
        "15_groups": lambda: f'{c("12_courses")}*COUNTA(\'07_divisions\'!A7:A{MAX_INPUT_ROWS})',
        "22_cohort_members": lambda: c("20_users_students"),
        "25_enrolments": lambda: c("15_groups"),
        "37_template_application": lambda: c("12_courses"),
        "51_academic_history": lambda: c("20_users_students"),
        "52_promotion_plan": lambda: c("20_users_students"),
        "54_next_year_courses": lambda: c("12_courses"),
        "55_next_year_cohorts": lambda: c("14_cohorts"),
        "56_next_year_groups": lambda: c("15_groups"),
        "57_next_year_enrolments": lambda: c("25_enrolments"),
        "assessment_plan": lambda: c("12_courses"),
        "attendance_policy": lambda: f"COUNTA('05_grades'!A7:A{MAX_INPUT_ROWS})",
        "course_certificates": lambda: c("12_courses"),
        "course_final_exams": lambda: c("12_courses"),
        "course_term_exams": lambda: (
            f'{c("12_courses")}*COUNTIFS(\'exam_terms\'!A7:A{MAX_INPUT_ROWS},"<>FINAL",'
            f'\'exam_terms\'!A7:A{MAX_INPUT_ROWS},"<>")'
        ),
        "gradebook_weights": lambda: c("12_courses"),
    }
    factory = formulas.get(sheet)
    return factory() if factory else None


def add_status_sheet(wb, year: str) -> None:
    if "status" in wb.sheetnames:
        del wb["status"]
    ws = wb.create_sheet("status", 2)
    ws.sheet_properties.tabColor = SYSTEM_TAB
    headers = ["type", "filename", "count", "status", "action"]
    ws.append(headers)
    control_rows = [
        (
            "automatic",
            "ALL_AUTOMATIC_SHEETS\nid_number_formula: rebuild every macro-generated sheet from current master rows",
            '=SUMIF(A7:A2000,"automatic",C7:C2000)',
            "CONTROL",
            macro_hyperlink("GenerateAllDerivedSheets", "Run all automatic macros"),
        ),
        (
            "automatic",
            "STATUS_ONLY\nid_number_formula: recalculate health formulas after manual review",
            "",
            "CONTROL",
            macro_hyperlink("RefreshStatus", "Refresh status"),
        ),
        (
            "automatic",
            "CLEAR_AUTOMATIC_SHEETS\nid_number_formula: remove data rows from macro-generated sheets",
            "",
            "CONTROL",
            macro_hyperlink("ClearAutomaticData", "Clear automatic data"),
        ),
        (
            "automatic",
            "RESET_AUTOMATIC_SHEETS\nid_number_formula: clear and regenerate all macro-generated sheets",
            "",
            "CONTROL",
            macro_hyperlink("ResetAutomaticData", "Reset and regenerate"),
        ),
        ("", "", "", "", ""),
    ]
    for row in control_rows:
        ws.append(row)

    row_by_sheet = {entry["sheet"]: index for index, entry in enumerate(SOURCE_FILES, start=7)}
    for entry in SOURCE_FILES:
        sheet = entry["sheet"]
        row_index = row_by_sheet[sheet]
        row_type = "automatic" if sheet in AUTOMATIC_SHEETS else "manual"
        source_csv = entry["source"].format(year=year, next_year=year_next(year))
        filename = (
            f"{entry['ordered']}\n"
            f"id_number_formula: {status_pattern(entry)}\n"
            f"source: {source_csv}"
        )
        count = f"=COUNTA({excel_quote(sheet)}!A7:A{MAX_INPUT_ROWS})"
        if row_type == "automatic":
            expected = status_expected_formula(sheet, row_by_sheet)
            if expected:
                status = f'=IF(C{row_index}=({expected}),"PASSED","FAILED")'
            else:
                status = f'=IF(C{row_index}>0,"PASSED","FAILED")'
            macro = AUTOMATIC_MACROS.get(sheet)
            action = macro_hyperlink(macro, "Run macro") if macro else "Automatic data"
        else:
            status = "MANUAL"
            action = "Manual edit"
        ws.append([row_type, filename, count, status, action])

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    widths = {"A": 16, "B": 78, "C": 18, "D": 16, "E": 28}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for cell in ws[1]:
        cell.fill = MANIFEST_FILL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=5):
        for cell in row:
            cell.fill = CONTROL_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        fill = REQUIRED_IF_USED_FILL if row[0].value == "automatic" else OPTIONAL_FILL
        for cell in row:
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.conditional_formatting.add(f"D7:D{ws.max_row}", FormulaRule(formula=['D7="PASSED"'], fill=PASS_FILL))
    ws.conditional_formatting.add(f"D7:D{ws.max_row}", FormulaRule(formula=['D7="FAILED"'], fill=FAIL_FILL))


def add_version_sheet(wb, year):
    ws = wb.create_sheet("_version")
    ws.sheet_properties.tabColor = SYSTEM_TAB
    rows = [
        ("Field", "Value"),
        ("template_version", TEMPLATE_VERSION),
        ("generated_on", date.today().isoformat()),
        ("academic_year", year),
        ("row_contract", "1 metadata, 2 guide, 3 summary, 4 formula/reference, 5 CSV header, 6 example, 7+ data"),
        ("max_input_rows", str(MAX_INPUT_ROWS)),
        ("moodle_target", "Moodle 5.2.x / 5.3-ready CSV process"),
        ("converter", "scripts/excel_to_source_csv.py"),
        ("report_generator", "scripts/generate_import_reports.py"),
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = MANIFEST_FILL
        cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    fit_columns(ws, max_width=80)


def add_sheet_index(wb, year):
    ws = wb.create_sheet("_sheet_index")
    ws.sheet_properties.tabColor = SYSTEM_TAB
    headers = ["sheet", "required", "standard_prefilled", "tab_color", "source_csv", "ordered_csv", "purpose"]
    ws.append(headers)
    for entry in SOURCE_FILES:
        tab = GENERATED_TAB if entry["sheet"] == "09_subject_matrix" else (REQUIRED_TAB if entry.get("required") else OPTIONAL_TAB)
        ws.append([
            entry["sheet"],
            "required" if entry.get("required") else "optional",
            "yes" if entry["sheet"] in STANDARD_PREFILL_SHEETS else "no",
            tab,
            entry["source"].format(year=year, next_year=year_next(year)),
            entry["ordered"],
            entry["purpose"],
        ])
    for cell in ws[1]:
        cell.fill = MANIFEST_FILL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
    for row in ws.iter_rows(min_row=2):
        fill = REQUIRED_IF_USED_FILL if row[1].value == "required" else OPTIONAL_FILL
        if row[0].value == "09_subject_matrix":
            fill = STAFF_FILL
        for cell in row:
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    fit_columns(ws, max_width=70)


def add_color_guide(wb):
    ws = wb.create_sheet("_color_guide", 1)
    ws.sheet_properties.tabColor = SYSTEM_TAB
    ws.append(["Workbook color guideline", "", "", ""])
    ws.append(["Area", "Color", "Meaning", "Operator action"])
    rows = [
        ("Header row", "Mandatory", "Required for import or reliable Moodle setup.", "Fill this before dry-run."),
        ("Header row", "Required if used", "Required when the optional sheet/feature is used.", "Fill it if you keep rows in that sheet."),
        ("Header row", "Optional", "Helpful metadata or optional Moodle setting.", "Fill only when available."),
        ("Guide row", "School-specific setup", "School identity, board, medium, grade, stream, division or subject setup.", "Use approved school codes."),
        ("Guide row", "Parent/family data", "Parent login, guardian relationship or parent-student link.", "Use verified parent details."),
        ("Guide row", "Student registration", "Student account, cohort or class/division registration.", "Use admission/roll data from school records."),
        ("Guide row", "Staff registration", "Teacher, principal, trustee or staff account data.", "Use official staff details."),
        ("Guide row", "Academic/course setup", "Courses, cohorts, groups, enrolments, certificates and assessment templates.", "Keep references aligned with course/cohort formulas."),
        ("Guide row", "Moodle/system reference", "Moodle roles, context identifiers, template references or operational metadata.", "Do not change unless the Moodle setup changes."),
    ]
    for row in rows:
        ws.append(row)

    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = HELP_FILL
    for cell in ws[2]:
        cell.fill = MANIFEST_FILL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER

    fill_by_label = {**REQUIREMENT_FILLS, **AUDIENCE_FILLS}
    for row in ws.iter_rows(min_row=3):
        label = row[1].value
        fill = fill_by_label.get(label, HELP_FILL)
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
    ws.freeze_panes = "A3"
    fit_columns(ws, max_width=64)


def add_data_sheet(wb, entry, year, source_root, sample_limit, include_example_row, lookup_names, prefill_standard_rows):
    ws = wb.create_sheet(entry["sheet"])
    ws.sheet_properties.tabColor = GENERATED_TAB if entry["sheet"] == "09_subject_matrix" else (
        REQUIRED_TAB if entry.get("required") else OPTIONAL_TAB
    )
    source_csv = entry["source"].format(year=year, next_year=year_next(year))
    headers = expected_headers(entry, year, source_root)
    guide_row = 2
    summary_row = 3
    pattern_row = 4
    header_row = 5
    example_row = 6 if include_example_row else None
    metadata = [
        "template_csv_file", source_csv,
        "ordered_csv_file", entry["ordered"],
        "purpose", entry["purpose"],
        "guide_row", str(guide_row),
        "summary_row", str(summary_row),
        "pattern_row", str(pattern_row),
        "header_row", str(header_row),
    ]
    if include_example_row:
        metadata.extend([
            "example_row", str(example_row),
            "example_note", "Row 6 shows concrete filled examples and is skipped during CSV generation",
        ])
    ws.append(metadata)
    guides = [column_guide(entry["sheet"], header, bool(entry.get("required"))) for header in headers]
    ws.append([f"{requirement} | {audience}" for requirement, audience in guides])
    summaries = [column_summary(header) for header in headers]
    patterns = [column_pattern(entry["sheet"], header) for header in headers]
    ws.append(summaries)
    ws.append(patterns)
    ws.append(headers)
    source_rows = read_rows(source_path(entry, year, source_root))
    if include_example_row:
        example = source_rows[0] if source_rows else {}
        ws.append([example.get(header, "") for header in headers])
    if prefill_standard_rows and entry["sheet"] in STANDARD_PREFILL_SHEETS:
        for row in source_rows:
            ws.append([row.get(header, "") for header in headers])
    if sample_limit:
        for row in source_rows[:sample_limit]:
            ws.append([row.get(header, "") for header in headers])
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=index)
        requirement, audience = guides[index - 1]
        summary = summaries[index - 1]
        pattern = patterns[index - 1]
        cell.comment = Comment(
            f"Requirement: {requirement}\n"
            f"Owner/use: {audience}\n"
            f"{summary}\n"
            f"Pattern/reference: {pattern}",
            "Master import",
        )
    first_input_row = (example_row + 1) if example_row else header_row + 1
    for index, header in enumerate(headers, start=1):
        lookup_key = LOOKUP_FIELDS.get(header)
        if not lookup_key and (header in BOOLEAN_FIELDS or header.startswith("is_") or header.startswith("allow_")):
            lookup_key = "boolean"
        if not lookup_key or lookup_key not in lookup_names:
            continue
        letter = get_column_letter(index)
        requirement = guides[index - 1][0]
        add_data_validation(
            ws,
            f"{letter}{first_input_row}:{letter}{MAX_INPUT_ROWS}",
            lookup_names[lookup_key],
            allow_blank=requirement == "Optional",
        )
    apply_conditional_formatting(ws, headers, guides, first_input_row)
    add_formula_helpers(ws, entry["sheet"], headers, guides, first_input_row)
    style_data_sheet(ws, guide_row, summary_row, pattern_row, header_row, example_row, guides)


def year_next(year):
    start, end = year.split("-")
    return f"{int(start) + 1}-{int(end) + 1}"


def build_workbook(
    path: Path,
    year: str,
    source_root: Path,
    sample_limit: int,
    include_example_row: bool,
    prefill_standard_rows: bool,
) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    add_dashboard(wb, year)
    add_manifest(wb, year)
    add_sheet_index(wb, year)
    add_version_sheet(wb, year)
    add_color_guide(wb)
    lookup_names = add_lookup_sheet(wb, build_lookup_values(year, source_root))
    for entry in SOURCE_FILES:
        add_data_sheet(wb, entry, year, source_root, sample_limit, include_example_row, lookup_names, prefill_standard_rows)
    add_status_sheet(wb, year)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.active = 0
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Create generic master import Excel workbooks.")
    parser.add_argument("--year", default="2026-2027")
    parser.add_argument("--source-root", default=str(PACK_ROOT))
    parser.add_argument("--template", default=str(PROCESS_ROOT / "templates" / "school_master_import_template.xlsx"))
    parser.add_argument("--sample", default=str(PROCESS_ROOT / "templates" / "sample_minimal_school_import.xlsx"))
    parser.add_argument("--sample-limit", type=int, default=3, help="Rows per sheet in the sample workbook.")
    args = parser.parse_args()

    source_root = Path(args.source_root).resolve()
    build_workbook(
        Path(args.template),
        args.year,
        source_root,
        sample_limit=0,
        include_example_row=True,
        prefill_standard_rows=True,
    )
    build_workbook(
        Path(args.sample),
        args.year,
        source_root,
        sample_limit=args.sample_limit,
        include_example_row=False,
        prefill_standard_rows=False,
    )
    print(f"Created template workbook: {args.template}")
    print(f"Created sample workbook: {args.sample}")


if __name__ == "__main__":
    main()
