#!/usr/bin/env python3
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

from common import PACK_ROOT, SOURCE_FILES, expected_headers, normalise_cell


def sheet_metadata(ws) -> dict[str, str]:
    cells = [normalise_cell(cell.value) for cell in ws[1]]
    metadata: dict[str, str] = {}
    for index in range(0, len(cells) - 1, 2):
        key = cells[index]
        value = cells[index + 1]
        if key:
            metadata[key] = value
    return metadata


def metadata_int(metadata: dict[str, str], key: str, default: int) -> int:
    value = metadata.get(key)
    if not value:
        return default
    try:
        return int(value)
    except ValueError:
        return default


def sheet_headers(ws) -> list[str]:
    metadata = sheet_metadata(ws)
    header_row = metadata_int(metadata, "header_row", 2)
    return [normalise_cell(cell.value) for cell in ws[header_row] if normalise_cell(cell.value)]


def validate(workbook: Path, year: str, source_root: Path) -> list[str]:
    errors: list[str] = []
    if not workbook.exists():
        return [f"Workbook does not exist: {workbook}"]

    wb = load_workbook(workbook, read_only=True, data_only=True)
    if "_manifest" not in wb.sheetnames:
        errors.append("Missing required _manifest sheet.")

    sheet_names = set(wb.sheetnames)
    for entry in SOURCE_FILES:
        sheet_name = entry["sheet"]
        if sheet_name not in sheet_names:
            if entry.get("required"):
                errors.append(f"Missing required sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        meta_a = normalise_cell(ws["A1"].value)
        meta_b = normalise_cell(ws["B1"].value)
        if meta_a != "template_csv_file":
            errors.append(f"{sheet_name}: row 1 must start with template_csv_file.")
        if not meta_b:
            errors.append(f"{sheet_name}: row 1 must include target CSV path in B1.")

        metadata = sheet_metadata(ws)
        header_row = metadata_int(metadata, "header_row", 2)
        guide_row = metadata_int(metadata, "guide_row", 0)
        summary_row = metadata_int(metadata, "summary_row", 0)
        pattern_row = metadata_int(metadata, "pattern_row", 0)
        if "header_row" in metadata and header_row < 2:
            errors.append(f"{sheet_name}: header_row metadata must be 2 or higher.")
        if guide_row and not any(normalise_cell(cell.value) for cell in ws[guide_row]):
            errors.append(f"{sheet_name}: guide_row {guide_row} is empty.")
        if summary_row and not any(normalise_cell(cell.value) for cell in ws[summary_row]):
            errors.append(f"{sheet_name}: summary_row {summary_row} is empty.")
        if pattern_row and not any(normalise_cell(cell.value) for cell in ws[pattern_row]):
            errors.append(f"{sheet_name}: pattern_row {pattern_row} is empty.")

        actual = sheet_headers(ws)
        expected = expected_headers(entry, year, source_root)
        if not actual:
            errors.append(f"{sheet_name}: header row {header_row} is empty.")
        elif expected and actual != expected:
            missing = [field for field in expected if field not in actual]
            extra = [field for field in actual if field not in expected]
            errors.append(f"{sheet_name}: header row {header_row} mismatch. missing={missing[:8]} extra={extra[:8]}")

    return errors


def main():
    parser = argparse.ArgumentParser(description="Validate the master import Excel workbook structure.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--year", default="2026-2027")
    parser.add_argument("--source-root", default=str(PACK_ROOT))
    args = parser.parse_args()

    errors = validate(Path(args.workbook), args.year, Path(args.source_root).resolve())
    if errors:
        print("Master Excel validation failed:")
        for error in errors[:80]:
            print(f"- {error}")
        if len(errors) > 80:
            print(f"... {len(errors) - 80} more errors")
        sys.exit(1)
    print("Master Excel validation passed.")


if __name__ == "__main__":
    main()
