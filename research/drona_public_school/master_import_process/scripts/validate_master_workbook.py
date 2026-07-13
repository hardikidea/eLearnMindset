#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook


PACK_ROOT = Path(__file__).resolve().parents[2]
REPO_ROOT = PACK_ROOT.parents[1]
MASTER_DIR = PACK_ROOT / "build" / "master_excel"
DEFAULT_XLSX = MASTER_DIR / "school_master_pack_2026_2027_full_predefined_master.xlsx"
DEFAULT_ODS = MASTER_DIR / "school_master_pack_2026_2027_full_predefined_master_macros.ods"
MACRO_SOURCE = PACK_ROOT / "master_import_process" / "scripts" / "libreoffice_master_tools.bas"
GENERATOR = PACK_ROOT / "master_import_process" / "scripts" / "create_libreoffice_macro_workbook.py"
README = MASTER_DIR / "README.md"
DEFAULT_SOFFICE = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
BUNDLED_SOFFICE = Path.home() / ".cache/codex-runtimes/codex-primary-runtime/dependencies/bin/override/soffice"
SOFFICE = Path(os.environ.get("SOFFICE", str(DEFAULT_SOFFICE)))


REQUIRED_SHEETS = [
    "_manifest",
    "_color_guide",
    "status",
    "_dashboard",
    "_sheet_index",
    "_version",
    "_lookups",
    "01_school_master",
    "02_academic_years",
    "03_boards",
    "04_mediums",
    "05_grades",
    "06_streams",
    "07_divisions",
    "08_subjects",
    "09_subject_matrix",
    "10_categories",
    "11_optional_categories",
    "12_courses",
    "13_courses_upload",
    "14_cohorts",
    "15_groups",
    "16_profile_fields",
    "17_custom_roles",
    "18_role_guidelines",
    "19_users_staff",
    "20_users_students",
    "21_users_parents",
    "22_cohort_members",
    "23_role_assignments",
    "24_parent_links",
    "25_enrolments",
    "26_lookup_values",
    "27_validation_rules",
    "28_source_refs",
    "29_summary",
    "30_master_template",
    "31_template_sections",
    "32_template_activities",
    "33_template_gradebook",
    "34_grade_band_adjust",
    "35_subject_adjust",
    "36_completion_defaults",
    "37_template_application",
    "38_template_custom_fields",
    "39_template_review",
    "40_certificate_policy",
    "41_report_access",
    "42_test_coverage",
    "43_content_template",
    "44_transition_models",
    "45_promotion_rules",
    "46_rollover_checklist",
    "47_promotion_policy",
    "48_promotion_status",
    "49_promotion_validation",
    "50_student_status",
    "51_academic_history",
    "52_promotion_plan",
    "53_promotion_actions",
    "54_next_year_courses",
    "55_next_year_cohorts",
    "56_next_year_groups",
    "57_next_year_enrolments",
    "58_alumni_cohorts",
    "59_archive_policy",
    "60_improvement_backlog",
    "61_compatibility",
    "assessment_plan",
    "attendance_policy",
    "course_certificates",
    "course_final_exams",
    "course_term_exams",
    "exam_terms",
    "gradebook_weights",
]


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def fail(errors: list[str], message: str) -> None:
    errors.append(message)


def check_file_exists(errors: list[str], path: Path) -> None:
    if not path.exists():
        fail(errors, f"missing file: {rel(path)}")
    else:
        print(f"OK {rel(path)}")


def check_required_sheets(errors: list[str], workbook) -> None:
    missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing:
        fail(errors, f"missing workbook sheets: {missing}")
        return
    print("OK required sheets exist")
    for sheet_name in REQUIRED_SHEETS:
        ws = workbook[sheet_name]
        if ws.max_row < 1 or ws.max_column < 1:
            fail(errors, f"{sheet_name}: empty sheet")
            continue
        headers = [
            ws.cell(1, col).value
            for col in range(1, ws.max_column + 1)
            if ws.cell(1, col).value not in (None, "")
        ]
        if not headers:
            fail(errors, f"{sheet_name}: missing row-1 headers")
    print("OK required sheet headers")


def check_status_sheet(errors: list[str], workbook) -> None:
    if "status" not in workbook.sheetnames:
        fail(errors, "missing workbook status sheet")
        return
    ws = workbook["status"]
    headers = [ws.cell(1, col).value for col in range(1, 6)]
    expected = ["type", "filename", "count", "status", "action"]
    if headers != expected:
        fail(errors, f"status sheet headers mismatch: {headers} != {expected}")
        return
    types = [str(ws.cell(row, 1).value or "") for row in range(7, ws.max_row + 1)]
    automatic = types.count("automatic")
    manual = types.count("manual")
    if automatic != 27:
        fail(errors, f"status sheet automatic row count mismatch: {automatic} != 27")
    if manual != 41:
        fail(errors, f"status sheet manual row count mismatch: {manual} != 41")
    actions = " ".join(str(ws.cell(row, 5).value or "") for row in range(2, min(ws.max_row, 12) + 1))
    if "vnd.sun.star.script" in actions or "Standard.MatrixTools" in actions:
        fail(errors, "xlsx status sheet must not contain document macro hyperlinks; use the .ods macro workbook")
    if "use the macro-enabled .ods workbook" not in actions:
        fail(errors, "xlsx status sheet missing macro-enabled .ods workbook guidance")
    print(f"OK status sheet: {automatic} automatic rows, {manual} manual rows")


def check_row_counts(errors: list[str], workbook) -> None:
    status_ws = workbook["status"]
    for row in range(2, status_ws.max_row + 1):
        filename = str(status_ws.cell(row, 2).value or "")
        if "course_term_exams.csv" in filename:
            formula = str(status_ws.cell(row, 4).value or "")
            if "SUMPRODUCT(COUNTIF('12_courses'!I7:I" not in formula:
                fail(errors, "status formula for course_term_exams must count 12_courses grade_code values")
            if "'05_grades'!E7:E" not in formula:
                fail(errors, "status formula for course_term_exams must use 05_grades moodle_label values")
            if "'exam_terms'!A7:A" in formula or "'exam_terms'!B7:B" in formula:
                fail(errors, "status formula for course_term_exams must not use a fixed exam_terms multiplier")
            print("OK course_term_exams status formula uses grade-driven term model")
            break
    else:
        fail(errors, "status sheet missing course_term_exams.csv row")


def check_macro_alignment(errors: list[str], workbook) -> None:
    macro_source = MACRO_SOURCE.read_text(encoding="utf-8")
    generator = GENERATOR.read_text(encoding="utf-8")
    readme = README.read_text(encoding="utf-8")

    macros_match = re.search(r'\["Individual macros", "([^"]+)"', generator)
    targets_match = re.search(r'\["Derived target sheets", "([^"]+)"', generator)
    if not macros_match or not targets_match:
        fail(errors, "macro guide rows missing in generator")
        return

    macros = macros_match.group(1).split(", ")
    targets = targets_match.group(1).split(", ")
    if len(macros) != len(targets):
        fail(errors, f"macro/target count mismatch: {len(macros)} != {len(targets)}")
    if len(macros) != 27:
        fail(errors, f"expected 27 individual macros, found {len(macros)}")

    missing_subs = [macro for macro in macros if f"Sub {macro}" not in macro_source]
    missing_docs = [macro for macro in macros if f"`{macro}`" not in readme]
    missing_target_docs = [target for target in targets if target not in readme]
    if missing_subs:
        fail(errors, f"missing macro Sub definitions: {missing_subs}")
    if missing_docs:
        fail(errors, f"missing README macro entries: {missing_docs}")
    if missing_target_docs:
        fail(errors, f"missing README target entries: {missing_target_docs}")

    sheet_refs = sorted(set(re.findall(r'RequireSheet\(oDoc, "([^"]+)"\)', macro_source)))
    missing_refs = [sheet for sheet in sheet_refs if sheet not in workbook.sheetnames]
    if missing_refs:
        fail(errors, f"Basic RequireSheet refs missing from workbook: {missing_refs}")
    else:
        print(f"OK macro guide/source/docs aligned; RequireSheet refs: {len(sheet_refs)}")


def check_ods_package(errors: list[str], ods_path: Path) -> None:
    required_entries = [
        "mimetype",
        "content.xml",
        "META-INF/manifest.xml",
        "Basic/Standard/MatrixTools.xml",
        "Basic/Standard/script-lb.xml",
        "Basic/script-lc.xml",
    ]
    with zipfile.ZipFile(ods_path) as archive:
        names = set(archive.namelist())
        missing_entries = [entry for entry in required_entries if entry not in names]
        if missing_entries:
            fail(errors, f"ODS missing package entries: {missing_entries}")
            return
        macro = archive.read("Basic/Standard/MatrixTools.xml").decode("utf-8")
        content = archive.read("content.xml").decode("utf-8", errors="ignore")
        for token in [
            "GenerateAllDerivedSheets",
            "RefreshStatus",
            "ClearAutomaticData",
            "ResetAutomaticData",
            "GenerateSummary",
            "GenerateRolloverChecklist",
            "GenerateOptionalYearCategoryModel",
            "GenerateStudentAcademicHistory",
            "GenerateStudentPromotionPlan",
            "GenerateAlumniCohorts",
            "GenerateArchivePolicy",
            "GenerateCompatibilityMatrix",
            "GenerateAssessmentPlan",
            "GenerateAttendancePolicy",
        ]:
            if token not in macro:
                fail(errors, f"ODS macro package missing {token}")
        if "00_MACRO_GUIDE" not in content:
            fail(errors, "ODS content missing 00_MACRO_GUIDE sheet")
        if "status" not in content:
            fail(errors, "ODS content missing status sheet")
        for token in [
            "vnd.sun.star.script",
            "GenerateAllDerivedSheets",
            "RefreshStatus",
            "ClearAutomaticData",
            "ResetAutomaticData",
        ]:
            if token not in content:
                fail(errors, f"ODS status sheet missing macro action token: {token}")
        if "SUMPRODUCT(COUNTIF([$&apos;12_courses&apos;.I6:.I100000]" not in content and "SUMPRODUCT(COUNTIF([$12_courses.I6:.I100000]" not in content:
            fail(errors, "ODS course_term_exams status formula must count 12_courses grade_code values from row 6")
        if "&apos;05_grades&apos;.E6" not in content and "$05_grades.E6" not in content:
            fail(errors, "ODS course_term_exams status formula must use 05_grades moodle_label values from row 6")
        if "00_MACRO_GUIDE" in content and "status" in content:
            print("OK ODS macro package")


def check_libreoffice_convert(errors: list[str], ods_path: Path) -> None:
    soffice = SOFFICE if SOFFICE.exists() else BUNDLED_SOFFICE
    if not soffice.exists():
        print(f"SKIP LibreOffice convert smoke; soffice not found at {SOFFICE} or {BUNDLED_SOFFICE}")
        return
    with tempfile.TemporaryDirectory(prefix="dps_master_smoke_") as tmp:
        tmpdir = Path(tmp)
        profile = tmpdir / "profile"
        command = [
            str(soffice),
            f"-env:UserInstallation=file://{profile}",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(tmpdir),
            str(ods_path),
        ]
        result = subprocess.run(command, capture_output=True, text=True)
        if result.returncode != 0:
            fail(errors, f"LibreOffice convert failed: {result.stderr or result.stdout}")
            return
        output = tmpdir / f"{ods_path.stem}.xlsx"
        if not output.exists():
            fail(errors, "LibreOffice convert did not create an XLSX output")
        else:
            print(f"OK LibreOffice convert smoke: {output.stat().st_size:,} bytes")


def main() -> int:
    parser = argparse.ArgumentParser(description="Smoke validate the school master-pack workbook and macro ODS.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--ods", type=Path, default=DEFAULT_ODS)
    parser.add_argument("--skip-libreoffice", action="store_true")
    args = parser.parse_args()

    errors: list[str] = []
    print("School master-pack workbook smoke validation")
    print("---------------------------------------")
    for path in [args.xlsx, args.ods, MACRO_SOURCE, GENERATOR, README]:
        check_file_exists(errors, path)
    if errors:
        for error in errors:
            print(f"ERROR {error}")
        return 1

    workbook = load_workbook(args.xlsx, read_only=False, data_only=False)
    print(f"OK workbook opens: {len(workbook.sheetnames)} sheets")
    if len(workbook.sheetnames) != 75:
        fail(errors, f"expected 75 sheets, found {len(workbook.sheetnames)}")
    if len(workbook.sheetnames) != len(set(workbook.sheetnames)):
        fail(errors, "duplicate sheet names found")
    if any(len(name) > 31 for name in workbook.sheetnames):
        fail(errors, "one or more sheet names exceed Excel's 31-character limit")
    else:
        print("OK sheet-name constraints")

    check_required_sheets(errors, workbook)
    check_status_sheet(errors, workbook)
    check_row_counts(errors, workbook)
    check_macro_alignment(errors, workbook)
    check_ods_package(errors, args.ods)
    if not args.skip_libreoffice:
        check_libreoffice_convert(errors, args.ods)

    if errors:
        print(f"Errors: {len(errors)}")
        for error in errors:
            print(f"ERROR {error}")
        return 1
    print("SMOKE PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
